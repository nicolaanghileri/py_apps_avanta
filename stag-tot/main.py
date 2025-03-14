import os
import glob
import openpyxl
from pprint import pprint


def collect_all_kontos(files):
    """
    Lädt alle Kontonummern (und optional den Namen in Spalte 2) aus
    allen übergebenen XLSX-Dateien und sammelt sie in einem Dict:
      {
        konto_nr_1: name_oder_none,
        konto_nr_2: name_oder_none,
        ...
      }
    """
    kontos_dict = {}

    for file_path in files:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True, min_row=2, max_row=ws.max_row-2):
            if not row or row[0] is None:  # leere oder ungültige Zeile
                continue

            konto_nr = row[0]
            # Falls du den Namen wirklich aus jeder Datei sammeln willst,
            # nimmst du hier z.B. das erste Vorkommen pro Konto:
            name = row[1] if len(row) > 1 else None

            if konto_nr not in kontos_dict:
                kontos_dict[konto_nr] = name

    return kontos_dict


def create_first_map(kontos_dict):
    """
    Erzeugt das `first_map`-Gerüst aus allen Kontonummern,
    wobei Index:
      0 -> Kontonummer
      1 -> Name (falls vorhanden, sonst None)
      2 -> Spalte für 061
      3 -> Spalte für 486
      4 -> Spalte für 495
      5 -> Spalte für 725
      6 -> Gesamtsumme
    """
    first_map = []
    for konto_nr, name in kontos_dict.items():
        first_map.append([konto_nr, name, 0.0, 0.0, 0.0, 0.0, 0.0])
    return first_map


def populate_totals(file_path, pos, first_map):
    """
    Liest das gegebene File und trägt die Salden (Spalte 4 im Excel) in die
    entsprechende Spalte (pos) der jeweiligen Kontonummer in `first_map` ein.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Index-Suche per Dictionary für schnellen Zugriff
    # key: konto_nr -> value: referenz auf row-list
    konto_dict = {row[0]: row for row in first_map}

    for row in ws.iter_rows(values_only=True, min_row=2,max_row=ws.max_row-2):
        if not row or row[0] is None:
            continue

        konto_nr = row[0]
        saldo = row[3]  # Annahme: Spalte 4 enthält den Saldo

        if konto_nr in konto_dict:
            konto_dict[konto_nr][pos] = saldo
        else:
            print(f"Warnung: Konto {konto_nr} in {file_path} nicht im Konto-Bestand!")


def process_files(source_dir):
    """
    - Sammelt zuerst die Kontonummern (und optional Namen) aus allen XLSX-Dateien im Verzeichnis.
    - Erstellt ein `first_map`, das alle Konten enthält.
    - Befüllt die Saldowerte pro Datei (anhand des Präfixes).
    - Berechnet anschließend die Summe in der letzten Spalte.
    """
    files = glob.glob(os.path.join(source_dir, "*.xlsx"))
    if not files:
        print("Keine Dateien gefunden!")
        return []

    # 1) Alle Kontos aus allen Dateien sammeln
    all_kontos_dict = collect_all_kontos(files)

    # 2) Grundgerüst auf Basis aller Kontos erstellen
    first_map = create_first_map(all_kontos_dict)

    # 3) Saldowerte je nach Präfix in `first_map` füllen
    for file_path in files:
        filename = os.path.basename(file_path)
        prefix = filename.split()[0]  # z.B. "061", "486", etc.

        if prefix == "061":
            populate_totals(file_path, 2, first_map)
        elif prefix == "486":
            populate_totals(file_path, 3, first_map)
        elif prefix == "495":
            populate_totals(file_path, 4, first_map)
        elif prefix == "725":
            populate_totals(file_path, 5, first_map)
        else:
            print(f"Unbekanntes Präfix '{prefix}' in Datei {filename}, wird übersprungen.")

    # 4) Gesamtsumme berechnen (letzte Spalte, Index 6)
    for konto in first_map:
        # Summiere alle Nicht-None-Werte aus den Spalten 2..5
        konto[6] = sum(x for x in konto[2:6] if x is not None)

    return first_map


def load_konto_zuordnung(filepath, sheet_name=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Falls das Sheet nicht angegeben ist, nimm das aktive
    ws = wb[sheet_name] if sheet_name else wb.active

    konto_map = {}

    # Wir gehen ab Zeile 2 durch (weil Zeile 1 Überschriften sind)
    for row in ws.iter_rows(min_row=5, values_only=True):
        # row[0] -> lokales Konto
        # row[1] -> lokale Bezeichnung
        # row[2] -> Konto Konzern
        # row[3] -> Bezeichnung Konto Konzern
        if not row or row[0] is None or row == '':
            # Wenn die Zeile leer ist oder kein lokales Konto enthält, überspringen
            continue

        lokales_konto = row[0]
        lokales_konto_name = row[1]
        konzern_konto = row[2]
        konzern_konto_name = row[3]

        if isinstance(lokales_konto, str):
            try:
                lokales_konto = int(lokales_konto)
            except ValueError:
                continue

        # Hier kannst du sicherstellen, dass es Zahlen sind
        # oder ggf. Strings. Je nach Formatierung in Excel.
        # Beispiel:

        konto_map[lokales_konto] = {
            "lokal_name": lokales_konto_name,
            "konzern_konto": konzern_konto,
            "konzern_name": konzern_konto_name
        }

    return konto_map


def attach_konzern_info(result_map, konto_map):
    """
    Ergänzt in jeder Zeile von `result_map` (der Form:
       [lok_kontonummer, lok_name, 61, 486, 495, 725, total]
    )
    die Konzerninformationen (kto, name).
    """
    for row in result_map:
        lokales_konto = row[0]  # z.B. 4007
        if lokales_konto in konto_map:
            row.append(konto_map[lokales_konto]["konzern_konto"])
            row.append(konto_map[lokales_konto]["konzern_name"])
        else:
            # Kein Mapping gefunden
            row.append(None)
            row.append(None)
    return result_map


def export_result_map(result_map, output_file):
    """
    Exportiert result_map in ein Excel-Dokument mit folgendem Header:
    Kontoname | Kontonummer | STAG - 061 | ICZH - 486 | ICGE - 495 | Zleep - 725 | Total | Kontoname Konzern | Kontonummer Konzern
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ergebnis"

    # Definiere den Header
    header = [
        "Kontoname",
        "Kontonummer",
        "STAG - 061",
        "ICZH - 486",
        "ICGE - 495",
        "Zleep - 725",
        "Total",
        "Kontoname Konzern",
        "Kontonummer Konzern"
    ]
    ws.append(header)

    # Füge die Zeilen aus result_map hinzu
    for row in result_map:
        ws.append(row)

    # Speichern
    wb.save(output_file)
    print(f"Export abgeschlossen: {output_file}")

if __name__ == "__main__":
    source_directory = os.path.abspath("../../source_tot/")
    
    konzern_map = load_konto_zuordnung("stag_kontenplan.xlsx")

    pprint(konzern_map)

    result_map = process_files(source_directory)
    result_map = attach_konzern_info(result_map, konzern_map)

    print(result_map)

    export_result_map(result_map, "ergebnis.xlsx")
    # Ausgabe zur Kontrolle

