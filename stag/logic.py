import os
import openpyxl

doc_number = 1

def validate_totals_before_anything(source_dir, label_status=None):
    """
    Prüft in allen Excel-Dateien im Ordner 'source_dir':
      - Wenn Kontonummer 2 leer ist (None oder Leerstring)
        UND der Wert in Spalte 'Total' != 0
        => FEHLER: Wir brechen ab und werfen eine Exception.

    label_status ist optional und dient zur GUI-Statusausgabe.
    """
    # Alle Excel-Dateien im Quellverzeichnis
    files = [
        f for f in os.listdir(source_dir)
        if os.path.isfile(os.path.join(source_dir, f)) and f.endswith(".xlsx")
    ]

    for f in files:
        file_path = os.path.join(source_dir, f)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Header in Zeile 1 auslesen
        headers = [cell.value for cell in ws[1]]

        # Versuchen, die Indizes für 'Kontonummer 2' und 'Total' zu finden
        try:
            col_konto2 = headers.index("Kontonummer 2")
            col_total = headers.index("Total")
        except ValueError:
            # Falls eine Spalte nicht existiert, gehen wir zur nächsten Datei
            wb.close()
            continue

        # Zeilen ab Zeile 5 validieren
        for row in ws.iter_rows(min_row=5, values_only=True):
            konto_value = row[col_konto2]
            total_value = row[col_total] if row[col_total] is not None else 0

            # Check: Kontonummer 2 leer + Total != 0 => Fehler
            if (konto_value is None or str(konto_value).strip() == "") and total_value != 0:
                msg = (
                    f"FEHLER in Datei '{f}': Kontonummer 2 ist leer, "
                    f"aber Total={total_value}. Abbruch!"
                )
                print(msg)
                wb.close()
                if label_status:
                    label_status.config(text=msg, fg="red")
                # Exception werfen, damit das Programm abbricht
                raise ValueError(msg)

        wb.close()

    print("Validierung erfolgreich: Keine fehlerhaften Zeilen gefunden.")
    if label_status:
        label_status.config(text="Validierung OK", fg="green")


def pivot_kst_columns(source_file, prefix):
    """
    Liest eine Excel-Datei und wandelt Zeilen aus KST-Spalten in ein Standardformat um.
    Gibt dabei ZWEI Werte pro Zeile zurück:
       1) Die reine KST-Nummer (z. B. '100')
       2) new_kst = prefix + cost_center_num (z. B. '486100')
    """
    global doc_number

    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source.active

    # Header-Zeile
    headers = [cell.value for cell in ws_source[1]]

    # Indizes der notwendigen Spalten
    try:
        col_kto2 = headers.index("Kontonummer 2")
        col_bez2 = headers.index("Bezeichnung 2")
    except ValueError as e:
        print(f"Fehlender Header in {source_file}: {e}")
        wb_source.close()
        return []

    # Finde alle KST-Spalten (Spalten, deren Header mit "KST" beginnt)
    kst_cols = [i for i, h in enumerate(headers) if h and str(h).startswith("KST")]

    rows_to_append = []

    # Zeilen ab Zeile 5
    for row in ws_source.iter_rows(min_row=5, values_only=True):
        kontonummer2 = row[col_kto2]
        bezeichnung2 = row[col_bez2]

        for c_idx in kst_cols:
            amount = row[c_idx]
            if amount and amount != 0:
                # KST-Nummer extrahieren, z.B. "KST 100" -> "100"
                cost_center_num = headers[c_idx].split()[1]
                new_kst = prefix + cost_center_num  # z. B. "486100"

                rows_to_append.append([
                    "01.01.2025",  # Value Date
                    "-",           # Buchungsart
                    prefix,        # Company number
                    cost_center_num,  # KST (reine Nummer)
                    new_kst,       # Cost Center (Prefix+Nummer)
                    kontonummer2,  # Cost type
                    bezeichnung2,  # Cost type description
                    "CHF",         # Currency
                    float(amount), # Amount
                    "01.01.2025",  # Entry date
                    "",            # Local Fibu Account
                    doc_number,    # Document Number
                    ""             # Document Description
                ])
                doc_number += 1

    wb_source.close()
    return rows_to_append


def revenue(source_file, prefix):
    """
    Ruft pivot_kst_columns auf und filtert die Zeilen nach bestimmten Kontonummern (Cost type).
    """
    # Erlaubte Kontonummern
    data = {
        "061": ["1725", "1580", "5615", "8080", "8040", "8720", "8643", "8640", "8650"],
        "486": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910","3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"],
        "495": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910","3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"],
        "725": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910","3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"]
    }

    # Alle Zeilen holen
    all_rows = pivot_kst_columns(source_file, prefix)

    # Falls prefix nicht in data: leere Liste
    allowed_accounts = data.get(prefix, [])

    # Filter: Nur wenn row[5] (Cost type) in den erlaubten Kontonummern ist
    filtered_rows = [row for row in all_rows if str(row[5]) in allowed_accounts]

    return filtered_rows


def save_revenue_to_excel(prefix_data_dict, output_directory):
    """
    Schreibt ein einzelnes 'Revenue_Report.xlsx' mit mehreren Sheets,
    pro Prefix ein Sheet, mit passendem Header (inkl. separater 'KST'-Spalte).
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    header = [
        "Value Date", "Buchungsart", "Company number", "KST", "Cost Center",
        "Cost type", "Cost type description", "Currency", "Amount",
        "Entry date", "Local Fibu Account", "Document Number", "Document Description"
    ]

    for prefix, rows in prefix_data_dict.items():
        ws = wb.create_sheet(title=prefix)
        ws.append(header)
        for row in rows:
            ws.append(row)

    output_file = os.path.join(output_directory, "Revenue_Report.xlsx")
    wb.save(output_file)
    print(f"[Revenue] Daten in '{output_file}' gespeichert.")


def checksum(source_file, min_row, column):
    """
    Summiert alle numerischen Werte (float/int) in einer bestimmten Spalte
    eines Excel-Files ab einer definierten Zeile.
    """
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source.active

    sum_value = 0.0
    for row in ws_source.iter_rows(min_row=min_row, values_only=True):
        value = row[column]
        if value is None:
            continue
        try:
            value_as_float = float(value)
            sum_value += value_as_float
        except ValueError:
            print(f"Warnung: '{value}' ist nicht numerisch und wird übersprungen.")

    wb_source.close()
    return sum_value


def floats_equal(a, b, epsilon=1e-9):
    return abs(a - b) < epsilon


def main_algo(source_dir, output_dir, selected_option, label_status=None):
    """
    - Prüft zuerst via validate_totals_before_anything(), ob Kontonummer 2
      leer + Total != 0 auftritt.
    - Monthly Report: Standard-Pivotisierung aller Files
    - Revenue: 1 Excel-File mit mehreren Sheets
    """
    # 1) Zuerst Validierung
    validate_totals_before_anything(source_dir, label_status)

    # 2) Ordner prüfen
    if not os.path.isdir(source_dir) or not os.path.isdir(output_dir):
        if label_status:
            label_status.config(text="Ungültige Pfade", fg="red")
        return

    # 3) Auswahl verarbeiten
    if selected_option == "Monthly Report":
        print("[Monthly Report] Starte Verarbeitung...")
        files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx")]

        final_rows = []
        final_checksum_files = 0

        for f in files:
            file_path = os.path.join(source_dir, f)
            prefix = f.split()[0]
            final_checksum_files += checksum(file_path, 5, 4)
            rows = pivot_kst_columns(file_path, prefix)
            if rows:
                final_rows.extend(rows)

        wb_pivot = openpyxl.Workbook()
        ws_pivot = wb_pivot.active
        ws_pivot.title = "Pivot_Gesamt"

        header = [
            "Value Date", "Buchungsart", "Company number", "Cost Center",
            "Cost type", "Cost type description", "Currency", "Amount",
            "Entry date", "Local Fibu Account", "Document Number", "Document Description"
        ]
        ws_pivot.append(header)

        for row in final_rows:
            ws_pivot.append(row)

        output_xlsx = os.path.join(output_dir, "MonthlyReport_STAG.xlsx")
        wb_pivot.save(output_xlsx)

        if label_status:
            label_status.config(text="Monthly Report erstellt!", fg="green")

    elif selected_option == "Revenue":
        print("[Revenue] Starte Verarbeitung...")
        files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx")]

        prefix_data_dict = {}
        for f in files:
            file_path = os.path.join(source_dir, f)
            prefix = f.split()[0]
            rev_data = revenue(file_path, prefix)
            if rev_data:
                prefix_data_dict.setdefault(prefix, []).extend(rev_data)

        if prefix_data_dict:
            save_revenue_to_excel(prefix_data_dict, output_dir)
            if label_status:
                label_status.config(text="Revenue Report erstellt!", fg="green")
        else:
            if label_status:
                label_status.config(text="Keine Revenue-Daten gefunden!", fg="red")
            print("Keine Revenue-Daten gefunden.")
    else:
        print(f"Unbekannte Option: {selected_option}")
        if label_status:
            label_status.config(text="Unbekannte Option!", fg="red")
