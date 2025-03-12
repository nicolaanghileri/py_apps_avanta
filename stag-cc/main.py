import os
import re
import time
import openpyxl
from datetime import datetime
from openpyxl import Workbook

record_number = 1

def record(date, konto1, konto2, text1, betrag, text2,ws,flag):
    global record_number

    buchungsebene = 0
    mwst_code = ""
    mwst_incl = ""
    mwst_land = ""
    mwst_koeff = ""
    mwst_konto = 0
    mwst_gkonto = 0
    mwst_sh = 0
    mwst_typ = 0

    # Datum als datetime-Objekt konvertieren
    if isinstance(date, str):
        try:
            date = datetime.strptime(date, '%d.%m.%Y')
        except ValueError:
            print("Ungültiges Datum:", date)
            return

    # Betrag als float konvertieren
    if isinstance(betrag, str):
        try:
            betrag = float(betrag.replace(',', '.'))
        except ValueError:
            print("Ungültiger Betrag:", betrag)
            return

    # MWST-Einstellungen basierend auf Flag
    if not flag:
        buchungsebene = 534
        mwst_code = 200
        mwst_incl = "I"
        mwst_land = "CH"
        mwst_koeff = 100
        mwst_konto = 61002
        mwst_gkonto = 28006
        mwst_sh = 2
        mwst_typ = 2

    # Zeile erstellen (eine flache Liste, keine verschachtelte Liste)
    row_to_append = [
        record_number, "J", date, konto1, konto2, text1, betrag, text2, "S",
        buchungsebene, 0, "", 0, "", "", "E", "", "", 0, "", "4221", "CHF", "CHF", 0, 0, 0, 0, 0, 0,
        "", "", "", "", mwst_code, 0, mwst_incl, 1, mwst_land, mwst_koeff, mwst_konto, mwst_gkonto, mwst_sh,
        0, 0, 0, 0, 0, mwst_typ, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, "", 0, 0
    ]

    # Zeile hinzufügen
    ws.append(row_to_append)
    record_number+=1

if __name__ == "__main__":
    start_time = time.time()
    file_path = os.path.join("../../source/486-Buchungen.xlsx")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active



    header = [
        "Record Nr", "Version", "Datum", "Konto", "Gegenkonto", "Text1", "Betrag", "Text2", "SH", 
        "Buchungsebene 1", "Kurs", "Kursart", "FWBetrag", "Sammelbuchungs Identifier", "Spec1", 
        "Applikationsidentifikation", "Reserve", "Valuta Datum", "Sam Position", "Reserve", "Mandant Nr", 
        "ISO", "ISO2", "Menge", "Ansatz", "Buchungsebene 2", "Gegenseite Buchungsebene 2", "Fond1", 
        "Fond2", "Reserve", "Reserve", "Reserve", "Codefeld", "Mwst Code", "Mwst Satz", "Mwst Incl", 
        "Mwst Methode", "Mwst Land", "Mwst Koeff", "Mwst Konto", "Mwst Gegenkonto", "Mwst SH", 
        "Mwst Betrag", "Mwst FW Betrag", "Mwst Betrag Rest", "Mwst FW Betrag Rest", "Reserve", "Mwst Typ", 
        "Reserve", "Reserve", "Reserve", "Geschäftsbereich", "Soll Ist", "HabenVerdSamBetrag", 
        "HabenVerdSamBetragFW", "Euro Koeff1", "Euro Koeff2", "Intercompany", "Kurs2", 
        "Konsolidierungscode", "Buchungsebene 3", "Gegenseite Buchungsebene 3"
    ]

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Buchungen Kreditkarten"

    output_ws.append(header)


    for row in ws.iter_rows(values_only=True,min_row=10):
        text = row[1]
        pattern_swisscard = r'/GR(\d+\.\d+)/DI-(\d+\.\d+)'
        pattern_payone = r"ALL\s+(\d+)\.\d+/NR\.\d+\nKOM\.\s+(\d+\.\d+)/DAT\.\d{2}\.\d{2}\.\d{4}"
        pattern_nexi = r"BRUTTO(\d+,\d+):KOM(\d+,\d+)"

        if re.search(pattern_swisscard, text):
            print("Detected: Swisscard AECS GmbH pattern")
            match = re.search(pattern_swisscard,text)
            record(date=row[0],konto1=16220,konto2=24011, text1="Swisscard AECS GmbH", betrag=match.group(1),text2="",ws=output_ws,flag=True)
            record(date=row[0],konto1=61002,konto2=16220, text1="Swisscard AECS GmbH", betrag=match.group(2),text2="Kommission",ws=output_ws,flag=False)
        elif re.search(pattern_payone, text):
            print("Detected: Payone GmbH pattern")
            match = re.search(pattern_payone,text)
            record(date=row[0],konto1=16220,konto2=24016, text1="Payone GmbH", betrag=match.group(1),text2="",ws=output_ws,flag=True)
            record(date=row[0],konto1=61002,konto2=16220, text1="Payone GmbH", betrag=match.group(2),text2="Kommission",ws=output_ws,flag=False)
        elif re.search(pattern_nexi, text):
            print("Detected: Nexi Germany GmbH pattern")
            match = re.search(pattern_nexi,text)
            record(date=row[0],konto1=16220,konto2=24016, text1="Nexi Germany GmbH", betrag=match.group(1),text2="",ws=output_ws,flag=True)
            record(date=row[0],konto1=61002,konto2=16220, text1="Nexi Germany GmbH", betrag=match.group(2),text2="Kommission",ws=output_ws,flag=False)
        end_time = time.time() 


    output_wb.save("output.xlsx")
    print("Ausführungszeit:", end_time - start_time, "Sekunden")
