# logic_pl.py
import os
import openpyxl
from datetime import datetime, timedelta
from openpyxl import Workbook

doc_number = 1

def get_firstday_lastmonth():
    today = datetime.today().date()
    first_day_this_month = today.replace(day=1)
    last_day_last_month = first_day_this_month - timedelta(days=1)
    first_day_last_month = last_day_last_month.replace(day=1)
    return first_day_last_month.strftime('%d.%m.%Y')

def pivot_kst_columns(source_file, prefix):
    global doc_number
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source.active
    headers = [cell.value for cell in ws_source[1]]
    try:
        col_kto2 = headers.index("Kontonummer 2")
        col_bez2 = headers.index("Bezeichnung 2")
    except ValueError as e:
        print(f"Fehlender Header in {source_file}: {e}")
        wb_source.close()
        return []
    # Finde alle KST-Spalten (Header beginnt mit "KST")
    kst_cols = [i for i, h in enumerate(headers) if h and str(h).startswith("KST")]
    rows_to_append = []
    correct_date = get_firstday_lastmonth()
    for row in ws_source.iter_rows(min_row=5, values_only=True):
        kontonummer2 = row[col_kto2]
        bezeichnung2 = row[col_bez2]
        for c_idx in kst_cols:
            amount = row[c_idx]
            if amount and amount != 0:
                cost_center_num = headers[c_idx].split()[1]  # z.B. "KST 100" -> "100"
                new_kst = prefix + cost_center_num
                rows_to_append.append([
                    correct_date,    # Value Date
                    "-",             # Buchungsart
                    prefix,          # Company number
                    new_kst,         # Cost Center (Prefix+Nummer)
                    kontonummer2,    # Cost type
                    bezeichnung2,    # Cost type description
                    "CHF",           # Currency
                    float(amount),   # Amount
                    "01.01.2025",    # Entry date
                    "",              # Local Fibu Account
                    doc_number,      # Document Number
                    bezeichnung2     # Document Description (gleicher Wert wie Cost type description)
                ])
                doc_number += 1
    wb_source.close()
    return rows_to_append

def revenue(source_file, prefix):
    allowed_data = {
        "061": ["1725", "1580", "5615", "8080", "8040", "8720", "8643", "8640", "8650"],
        "486": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910", "3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"],
        "495": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910", "3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"],
        "725": ["1090", "1099", "1110", "1115", "1155", "1150", "1290", "3410", "1225", "8261",
                "1910", "3312", "1260", "1265", "1280", "2110", "2150", "2710", "7120", "8643", "8640"]
    }
    all_rows = pivot_kst_columns(source_file, prefix)
    filtered_rows = [row for row in all_rows if str(row[4]) in allowed_data.get(prefix, [])]
    return filtered_rows

def checksum(source_file, min_row, column):
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source.active
    sum_value = 0.0
    for row in ws_source.iter_rows(min_row=min_row, values_only=True):
        value = row[column]
        if value is None:
            continue
        try:
            sum_value += float(value)
        except ValueError:
            print(f"Warnung: '{value}' ist nicht numerisch und wird übersprungen.")
    wb_source.close()
    return sum_value

def validate_totals_before_anything(source_dir, label_status=None):
    files = [f for f in os.listdir(source_dir)
             if os.path.isfile(os.path.join(source_dir, f)) and f.endswith(".xlsx")]
    for f in files:
        file_path = os.path.join(source_dir, f)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        try:
            col_konto2 = headers.index("Kontonummer 2")
            col_total = headers.index("Total")
        except ValueError:
            wb.close()
            continue
        for row in ws.iter_rows(min_row=5, values_only=True):
            konto_value = row[col_konto2]
            total_value = row[col_total] if row[col_total] is not None else 0
            if (konto_value is None or str(konto_value).strip() == "") and total_value != 0:
                msg = f"FEHLER in Datei '{f}': Kontonummer 2 ist leer, aber Total={total_value}. Abbruch!"
                print(msg)
                wb.close()
                if label_status:
                    label_status.config(text=msg, fg="red")
                raise ValueError(msg)
        wb.close()
    print("Validierung erfolgreich: Keine fehlerhaften Zeilen gefunden.")
    if label_status:
        label_status.config(text="Validierung OK", fg="green")


def process_pl_report(source_dir, output_dir, label_status=None):
    global doc_number
    doc_number = 1  # Nach jedem Lauf zurücksetzen
    
    # Zuerst Validierung
    validate_totals_before_anything(source_dir, label_status)
    
    files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx")]

    # P&L-Report erstellen
    final_rows = []
    for f in files:
        file_path = os.path.join(source_dir, f)
        prefix = f.split()[0]
        final_rows.extend(pivot_kst_columns(file_path, prefix))

    # Excel-Workbook für das P&L-Resultat anlegen
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot_Gesamt"

    header = [
        "Value Date", "Buchungsart", "Company number", "Cost Center", "Cost type",
        "Cost type description", "Currency", "Amount", "Entry date", "Local Fibu Account",
        "Document Number", "Document Description"
    ]
    ws.append(header)

    for row in final_rows:
        ws.append(row)

    output_file = os.path.join(output_dir, "STAG-MonthlyReport.xlsx")
    wb.save(output_file)

    print("Monthly Report (P&L) erstellt!")
    if label_status:
        label_status.config(text="Monthly Report (P&L) erstellt!", fg="green")
        
        
def process_revenue_report(source_dir, output_dir, label_status=None):
    global doc_number
    doc_number = 1  # Nach jedem Lauf zurücksetzen

    validate_totals_before_anything(source_dir, label_status)

    files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx")]

    # Revenue-Report erstellen
    all_rows = []
    for f in files:
        file_path = os.path.join(source_dir, f)
        prefix = f.split()[0]
        rev_data = revenue(file_path, prefix)
        if rev_data:
            all_rows.extend(rev_data)

    wb_rev = Workbook()
    ws_rev = wb_rev.active
    ws_rev.title = "Revenue"

    header = [
        "Value Date", "Buchungsart", "Company number", "Cost Center", "Cost type",
        "Cost type description", "Currency", "Amount", "Entry date", "Local Fibu Account",
        "Document Number", "Document Description"
    ]
    ws_rev.append(header)

    for row in all_rows:
        ws_rev.append(row)

    output_file_rev = os.path.join(output_dir, "STAG-Revenue.xlsx")
    wb_rev.save(output_file_rev)

    print("Revenue Report erstellt!")
    if label_status:
        label_status.config(text="Revenue Report erstellt!", fg="green")