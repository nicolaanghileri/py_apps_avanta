# logic_cc.py
import os
import re
import time
import openpyxl
from openpyxl import Workbook


record_number = 1
mandant_nr = "0000"

def record(date, konto1, konto2, text1, betrag, text2, ws, flag):
    global record_number
    global mandant_nr
    
    buchungsebene = 0
    mwst_code = ""
    mwst_incl = ""
    mwst_land = ""
    mwst_koeff = 0
    mwst_konto = 0
    mwst_gkonto = 0
    mwst_sh = 0
    mwst_typ = 0
    if isinstance(betrag, str):
        try:
            betrag = float(betrag.replace(',', '.'))
        except ValueError:
            print("Ungültiger Betrag:", betrag)
            return
    if not flag:
        buchungsebene = 534
        mwst_code = 200
        mwst_incl = "I"
        mwst_land = "CH"
        mwst_koeff = 100
        mwst_konto = 61002
        mwst_gkonto = 28006
        mwst_sh = 2
        mwst_typ = 2
    row_to_append = [
        record_number, "J", date, konto1, konto2, text1, betrag, text2, "S",
        buchungsebene, 0, "", 0, "", 0, "", "", "E", "", "", 0, "", mandant_nr, "CHF", "CHF",
        0, 0, 0, 0, 0, 0, "", "", "", "", mwst_code, 0, mwst_incl, 1, mwst_land, mwst_koeff,
        mwst_konto, mwst_gkonto, mwst_sh, 0, 0, 0, 0, 0, mwst_typ, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, "", 0, 0
    ]
    ws.append(row_to_append)
    record_number += 1


def import_credit_card_payments(source_dir, output_dir, label_status=None):
    global record_number
    global mandant_nr
    
    record_number = 1
    start_time = time.time()
    files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx")]
    
    for file in files:
        
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "Buchungen Kreditkarten"
        header = [
            "Record Nr", "Version", "Datum", "Konto", "Gegenkonto", "Text1", "Betrag", "Text2", "SH", 
            "Buchungsebene 1", "Gegenseite Buchungsebene 1", "Belegnummer", "Kurs", "Kursart", "FWBetrag", 
            "Sammelbuchungs Identifier", "Spec1", "Applikationsidentifikation", "Reserve", "Valuta Datum", 
            "Sam Position", "Reserve", "Mandant Nr", "ISO", "ISO2", "Menge", "Ansatz", "Buchungsebene 2", 
            "Gegenseite Buchungsebene 2", "Fond1", "Fond2", "Reserve", "Reserve", "Reserve", "Codefeld", 
            "Mwst Code", "Mwst Satz", "Mwst Incl", "Mwst Methode", "Mwst Land", "Mwst Koeff", "Mwst Konto", 
            "Mwst Gegenkonto", "Mwst SH", "Mwst Betrag", "Mwst FW Betrag", "Mwst Betrag Rest", 
            "Mwst FW Betrag Rest", "Reserve", "Mwst Typ", "Reserve", "Reserve", "Reserve", "Geschäftsbereich", 
            "Soll Ist", "HabenVerdSamBetrag", "HabenVerdSamBetragFW", "Euro Koeff1", "Euro Koeff2", 
            "Intercompany", "Kurs2", "Konsolidierungscode", "Buchungsebene 3", "Gegenseite Buchungsebene 3"
        ]
        output_ws.append(header)
        
        file_path = os.path.join(source_dir, file)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        b5_value = ws['B5'].value or ""
        match_konto = re.search(r'\b507185-(11-\d)\b', b5_value)
        mandant_map = {
            "11-5": "4221", #ICZH
            "11-7": "4222", #ICGE
            "11-8": "4223"  #Zleep
        }
        
        if match_konto:
            suffix = match_konto.group(1)
            mandant_nr = mandant_map.get(suffix, "0000")
        
        for row in ws.iter_rows(values_only=True, min_row=10):
            text = row[1]
            pattern_swisscard = r'/GR(\d+\.\d+)/DI-(\d+\.\d+)'
            pattern_payone = r"ALL\s+(\d+)\.\d+/NR\.\d+\nKOM\.\s+(\d+\.\d+)/DAT\.\d{2}\.\d{2}\.\d{4}"
            pattern_nexi = r"BRUTTO(\d+,\d+):KOM(\d+,\d+)"
            if re.search(pattern_swisscard, text):
                match = re.search(pattern_swisscard, text)
                record(date=row[0], konto1=16220, konto2=24011, text1="Swisscard AECS GmbH",
                       betrag=match.group(1), text2="", ws=output_ws, flag=True)
                record(date=row[0], konto1=61002, konto2=16220, text1="Swisscard AECS GmbH",
                       betrag=match.group(2), text2="Kommission", ws=output_ws, flag=False)
            elif re.search(pattern_payone, text):
                match = re.search(pattern_payone, text)
                record(date=row[0], konto1=16220, konto2=24015, text1="Payone GmbH",
                       betrag=match.group(1), text2="", ws=output_ws, flag=True)
                record(date=row[0], konto1=61002, konto2=16220, text1="Payone GmbH",
                       betrag=match.group(2), text2="Kommission", ws=output_ws, flag=False)
            elif re.search(pattern_nexi, text):
                match = re.search(pattern_nexi, text)
                record(date=row[0], konto1=16220, konto2=24016, text1="Nexi Germany GmbH",
                       betrag=match.group(1), text2="", ws=output_ws, flag=True)
                record(date=row[0], konto1=61002, konto2=16220, text1="Nexi Germany GmbH",
                       betrag=match.group(2), text2="Kommission", ws=output_ws, flag=False)
                
            file_name = f"{mandant_nr}-CC-PaymentsImport.xlsx"
            output_file = os.path.join(output_dir, file_name)
            output_wb.save(output_file)
            wb.close()
        record_number = 1

    
    end_time = time.time()
    if label_status:
        label_status.config(text="Import Credit Card Payments abgeschlossen!", fg="green")
    print("Ausführungszeit:", end_time - start_time, "Sekunden")
