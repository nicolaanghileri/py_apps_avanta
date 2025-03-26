# logic_con.py
import os
import glob
import openpyxl
from openpyxl import Workbook

def collect_all_kontos(files):
    kontos_dict = {}
    for file_path in files:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True, min_row=2, max_row=ws.max_row-2):
            if not row or row[0] is None:
                continue
            konto_nr = row[0]
            name = row[1] if len(row) > 1 else None
            if konto_nr not in kontos_dict:
                kontos_dict[konto_nr] = name
        wb.close()
    return kontos_dict

def create_first_map(kontos_dict):
    first_map = []
    for konto_nr, name in kontos_dict.items():
        first_map.append([konto_nr, name, 0.0, 0.0, 0.0, 0.0, 0.0])
    return first_map

def populate_totals(file_path, pos, first_map):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    konto_dict = {row[0]: row for row in first_map}
    for row in ws.iter_rows(values_only=True, min_row=2, max_row=ws.max_row-2):
        if not row or row[0] is None:
            continue
        konto_nr = row[0]
        saldo = row[3]  # Annahme: Spalte 4 enthält den Saldo
        if konto_nr in konto_dict:
            konto_dict[konto_nr][pos] = saldo
        else:
            print(f"Warnung: Konto {konto_nr} in {file_path} nicht im Konto-Bestand!")
    wb.close()

def process_files(source_dir):
    files = glob.glob(os.path.join(source_dir, "*.xlsx"))
    if not files:
        print("Keine Dateien gefunden!")
        return []
    all_kontos_dict = collect_all_kontos(files)
    first_map = create_first_map(all_kontos_dict)
    for file_path in files:
        filename = os.path.basename(file_path)
        prefix = filename.split()[0]
        if prefix == "061":
            populate_totals(file_path, 2, first_map)
        elif prefix == "486":
            populate_totals(file_path, 3, first_map)
        elif prefix == "495":
            populate_totals(file_path, 4, first_map)
        elif prefix == "725":
            populate_totals(file_path, 5, first_map)
        else:
            print(f"Unbekanntes Präfix '{prefix}' in Datei {filename}, wird übersprungen.")
    for konto in first_map:
        konto[6] = sum(x for x in konto[2:6] if x is not None)
    return first_map

def load_konto_zuordnung(filepath, sheet_name=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    konto_map = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or row[0] is None or row == '':
            continue
        lokales_konto = row[0]
        lokales_konto_name = row[1]
        konzern_konto = row[2]
        konzern_konto_name = row[3]
        try:
            if isinstance(lokales_konto, str):
                lokales_konto = int(lokales_konto)
        except ValueError:
            continue
        konto_map[lokales_konto] = {
            "lokal_name": lokales_konto_name,
            "konzern_konto": konzern_konto,
            "konzern_name": konzern_konto_name
        }
    wb.close()
    return konto_map

def attach_konzern_info(result_map, konto_map):
    for row in result_map:
        lokales_konto = row[0]
        if lokales_konto in konto_map:
            row.append(konto_map[lokales_konto]["konzern_konto"])
            row.append(konto_map[lokales_konto]["konzern_name"])
        else:
            row.append(None)
            row.append(None)
    return result_map

def export_result_map(result_map, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ergebnis"
    header = ["Kontoname", "Kontonummer", "STAG - 061", "ICZH - 486", "ICGE - 495",
              "Zleep - 725", "Total", "Kontoname Konzern", "Kontonummer Konzern"]
    ws.append(header)
    for row in result_map:
        ws.append(row)
    wb.save(output_file)
    print(f"Export abgeschlossen: {output_file}")

def consolidation(source_dir, output_dir, label_status=None):
    mapping_file = os.path.join("", "stag_kontenplan.xlsx")
    if not os.path.exists(mapping_file):
        print("Kontozuordnung (stag_kontenplan.xlsx) nicht gefunden!")
        if label_status:
            label_status.config(text="Kontozuordnung nicht gefunden!", fg="red")
        return
    konto_map = load_konto_zuordnung(mapping_file)
    result_map = process_files(source_dir)
    result_map = attach_konzern_info(result_map, konto_map)
    output_file = os.path.join(output_dir, "ergebnis_consolidation.xlsx")
    export_result_map(result_map, output_file)
    if label_status:
        label_status.config(text="Consolidation abgeschlossen!", fg="green")
